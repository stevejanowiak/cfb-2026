"""Rebuildable engine: reads everything from the two files in /mnt/user-data.
No local state required, so this survives a sandbox reset."""
import openpyxl, math
from statistics import NormalDist
WB='/mnt/user-data/outputs/NCAA_2026_Win_Projections.xlsx'
HFA, SD = 2.5, 16
def load():
    wb=openpyxl.load_workbook(WB, data_only=True)
    rt={}
    R=wb['Ratings']
    for r in range(2,R.max_row+1):
        n=R.cell(r,1).value
        if n: rt[n]=R.cell(r,2).value
    S=wb['Schedule']; sched={}
    for r in range(2,S.max_row+1):
        t=S.cell(r,1).value
        if not t: continue
        gl=[]
        for k in range(13):
            b=2+k*5
            opp=S.cell(r,b+1).value
            if opp: gl.append((opp, S.cell(r,b+2).value))
        sched[t]=gl
    return rt, sched
def project(t, rt, sched, results):
    """results: dict team -> list of 1/0 by game index (chronological)."""
    res=results.get(t,[])
    W=sum(res); ps=[]
    for i,(opp,site) in enumerate(sched[t]):
        if i < len(res): continue
        adj = HFA if site=='Home' else (-HFA if site=='Away' else 0)
        ps.append(NormalDist(0,SD).cdf(rt[t]-rt[opp]+adj))
    D=[1.0]
    for p in ps:
        nd=[0.0]*(len(D)+1)
        for k,v in enumerate(D): nd[k]+=v*(1-p); nd[k+1]+=v*p
        D=nd
    return dict(wins=W, played=len(res), remain=len(ps), exp=W+sum(ps), dist=D)
def bet(t, side, line, payout, st):
    p_over=sum(v for k,v in enumerate(st['dist']) if k+st['wins']>line)
    win = p_over if side=='over' else 1-p_over
    gms = (st['exp']-line) if side=='over' else (line-st['exp'])
    magic = (math.ceil(line)-st['wins']) if side=='over' else (st['wins']+st['remain']-math.floor(line))
    return dict(win=win, gms=gms, magic=magic, value=win*payout)
SLATE=[("Florida St.","over",6.5,254.00),("Miami-OH","over",7.5,205.00),
 ("WVU","under",5.5,234.00),("Pittsburgh","under",7.5,225.00),("Utah","under",8.5,225.00),
 ("Alabama","under",8.5,205.00),("Wisconsin","under",6.5,186.96),("Arizona St.","under",6.5,195.24),
 ("Ole Miss","under",7.5,230.00),("MTSU","under",3.5,225.00)]
